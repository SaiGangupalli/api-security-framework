#!/usr/bin/env python3
"""
Test Case Processor - Main Script
Processes Excel test cases and rephrases them using AI
"""

import os
import pandas as pd
import time
from typing import List, Dict, Any
from test_case_ai import TestCaseAI
from config import Config
import logging

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/test_processor.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class TestCaseProcessor:
    def __init__(self):
        self.config = Config()
        self.ai_processor = TestCaseAI()
        
    def load_excel_file(self, file_path: str) -> pd.DataFrame:
        """Load Excel file and return DataFrame"""
        try:
            df = pd.read_excel(file_path)
            logger.info(f"Loaded {len(df)} test cases from {file_path}")
            return df
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            raise
    
    def validate_columns(self, df: pd.DataFrame) -> bool:
        """Validate required columns exist"""
        required_columns = self.config.REQUIRED_COLUMNS
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            logger.error(f"Missing required columns: {missing_columns}")
            return False
        
        logger.info("All required columns found")
        return True
    
    def process_batch(self, df: pd.DataFrame, start_idx: int = 0, batch_size: int = 10) -> pd.DataFrame:
        """Process a batch of test cases"""
        end_idx = min(start_idx + batch_size, len(df))
        batch_df = df.iloc[start_idx:end_idx].copy()
        
        logger.info(f"Processing batch: rows {start_idx} to {end_idx-1}")
        
        for idx, row in batch_df.iterrows():
            try:
                logger.info(f"Processing test case {idx + 1}")
                
                # Extract original data
                original_description = str(row.get('Test Case Description', ''))
                original_steps = str(row.get('Test Case Steps', ''))
                
                # Generate rephrased content using AI
                rephrased_data = self.ai_processor.rephrase_test_case(
                    description=original_description,
                    steps=original_steps
                )
                
                # Update the DataFrame
                batch_df.loc[idx, 'Test Case Description'] = rephrased_data['description']
                batch_df.loc[idx, 'Test Case Steps'] = rephrased_data['steps']
                batch_df.loc[idx, 'Expected Results'] = rephrased_data['expected_results']
                
                # Add processing metadata
                batch_df.loc[idx, 'AI_Processed'] = True
                batch_df.loc[idx, 'Processing_Timestamp'] = pd.Timestamp.now()
                
                # Rate limiting
                time.sleep(self.config.API_DELAY)
                
            except Exception as e:
                logger.error(f"Error processing row {idx}: {e}")
                batch_df.loc[idx, 'AI_Processed'] = False
                batch_df.loc[idx, 'Processing_Error'] = str(e)
        
        return batch_df
    
    def save_processed_data(self, df: pd.DataFrame, output_path: str):
        """Save processed data to Excel"""
        try:
            # Create output directory if it doesn't exist
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save to Excel
            df.to_excel(output_path, index=False)
            logger.info(f"Saved processed data to {output_path}")
            
        except Exception as e:
            logger.error(f"Error saving file: {e}")
            raise
    
    def process_file(self, input_file: str, output_file: str = None, batch_size: int = 10):
        """Main processing function"""
        try:
            # Load data
            df = self.load_excel_file(input_file)
            
            # Validate columns
            if not self.validate_columns(df):
                raise ValueError("Column validation failed")
            
            # Set output file name if not provided
            if output_file is None:
                base_name = os.path.splitext(os.path.basename(input_file))[0]
                output_file = f"output/{base_name}_rephrased.xlsx"
            
            # Process all data in batches
            processed_dfs = []
            total_rows = len(df)
            
            for start_idx in range(0, total_rows, batch_size):
                batch_df = self.process_batch(df, start_idx, batch_size)
                processed_dfs.append(batch_df)
                
                logger.info(f"Completed batch {start_idx//batch_size + 1}")
            
            # Combine all processed batches
            final_df = pd.concat(processed_dfs, ignore_index=True)
            
            # Save results
            self.save_processed_data(final_df, output_file)
            
            # Print summary
            successful_processes = final_df['AI_Processed'].sum()
            logger.info(f"Processing complete! {successful_processes}/{len(final_df)} test cases processed successfully")
            
            return final_df
            
        except Exception as e:
            logger.error(f"Processing failed: {e}")
            raise

def main():
    """Main execution function"""
    processor = TestCaseProcessor()
    
    # Example usage
    input_file = "input/test_cases.xlsx"
    
    try:
        result_df = processor.process_file(
            input_file=input_file,
            batch_size=10
        )
        print("✅ Processing completed successfully!")
        print(f"📊 Processed {len(result_df)} test cases")
        
    except Exception as e:
        print(f"❌ Processing failed: {e}")

if __name__ == "__main__":
    main()










"""
Test Case AI Processing Module
Handles AI-powered rephrasing of test cases using OpenAI API
"""

import openai
import json
import logging
from typing import Dict, Any
from config import Config

logger = logging.getLogger(__name__)

class TestCaseAI:
    def __init__(self):
        self.config = Config()
        self.client = openai.OpenAI(api_key=self.config.OPENAI_API_KEY)
    
    def create_prompt(self, description: str, steps: str) -> str:
        """Create a structured prompt for AI processing"""
        prompt = f"""
You are a QA expert tasked with rephrasing test cases to make them clearer and more professional while maintaining their original intent and coverage.

Original Test Case:
Description: {description}
Steps: {steps}

Please provide a JSON response with the following structure:
{{
    "description": "Rephrased test case description (clear, concise, professional)",
    "steps": "Rephrased test steps (numbered, clear, actionable)",
    "expected_results": "Expected results based on the description and steps (specific, measurable outcomes)"
}}

Guidelines:
1. Keep the original intent and scope of the test
2. Make language more professional and clear
3. Ensure steps are actionable and sequential
4. Expected results should be specific and verifiable
5. Use proper testing terminology
6. Maintain all critical test scenarios from the original

Return only the JSON response, no additional text.
"""
        return prompt
    
    def parse_ai_response(self, response_text: str) -> Dict[str, str]:
        """Parse AI response and extract structured data"""
        try:
            # Try to parse as JSON
            data = json.loads(response_text)
            
            # Validate required fields
            required_fields = ['description', 'steps', 'expected_results']
            for field in required_fields:
                if field not in data:
                    raise ValueError(f"Missing required field: {field}")
            
            return data
            
        except json.JSONDecodeError:
            logger.error("Failed to parse AI response as JSON")
            # Fallback: try to extract content manually
            return self.fallback_parse(response_text)
    
    def fallback_parse(self, response_text: str) -> Dict[str, str]:
        """Fallback parsing when JSON parsing fails"""
        logger.warning("Using fallback parsing method")
        
        # Simple fallback - return original with minimal changes
        return {
            "description": response_text[:200] + "..." if len(response_text) > 200 else response_text,
            "steps": "Steps need manual review - AI parsing failed",
            "expected_results": "Expected results need manual review - AI parsing failed"
        }
    
    def rephrase_test_case(self, description: str, steps: str) -> Dict[str, str]:
        """Main function to rephrase a test case using AI"""
        try:
            # Create prompt
            prompt = self.create_prompt(description, steps)
            
            # Call OpenAI API
            response = self.client.chat.completions.create(
                model=self.config.AI_MODEL,
                messages=[
                    {"role": "system", "content": "You are a professional QA expert specialized in test case documentation."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=self.config.MAX_TOKENS,
                temperature=self.config.TEMPERATURE
            )
            
            # Extract response text
            response_text = response.choices[0].message.content.strip()
            logger.debug(f"AI Response: {response_text}")
            
            # Parse response
            parsed_data = self.parse_ai_response(response_text)
            
            return parsed_data
            
        except Exception as e:
            logger.error(f"Error in AI processing: {e}")
            # Return fallback response
            return {
                "description": f"Error processing: {description}",
                "steps": f"Error processing: {steps}",
                "expected_results": f"Manual review required due to processing error: {str(e)}"
            }
    
    def test_connection(self) -> bool:
        """Test AI API connection"""
        try:
            response = self.client.chat.completions.create(
                model=self.config.AI_MODEL,
                messages=[{"role": "user", "content": "Hello, are you working?"}],
                max_tokens=10
            )
            logger.info("AI API connection successful")
            return True
        except Exception as e:
            logger.error(f"AI API connection failed: {e}")
            return False

# Alternative implementation using local AI models (if you prefer not to use OpenAI)
class LocalTestCaseAI:
    """Alternative implementation using local models like Ollama"""
    
    def __init__(self):
        self.config = Config()
    
    def rephrase_test_case(self, description: str, steps: str) -> Dict[str, str]:
        """Rephrase using local AI model"""
        try:
            import requests
            
            prompt = f"""
Rephrase this test case to be more professional and clear:

Description: {description}
Steps: {steps}

Provide:
1. Improved description
2. Clear, numbered steps
3. Expected results
"""
            
            # Example for Ollama local API
            response = requests.post(
                'http://localhost:11434/api/generate',
                json={
                    'model': 'llama2',  # or your preferred local model
                    'prompt': prompt,
                    'stream': False
                }
            )
            
            if response.status_code == 200:
                ai_response = response.json()['response']
                # Parse the response (you'd need to implement parsing logic)
                return self.parse_local_response(ai_response)
            else:
                raise Exception(f"Local AI API error: {response.status_code}")
                
        except Exception as e:
            logger.error(f"Local AI processing error: {e}")
            return {
                "description": description,
                "steps": steps,
                "expected_results": "Expected results need manual definition"
            }
    
    def parse_local_response(self, response: str) -> Dict[str, str]:
        """Parse local AI response"""
        # Implement parsing logic based on your local model's response format
        lines = response.split('\n')
        return {
            "description": "Rephrased description (implement parsing)",
            "steps": "Rephrased steps (implement parsing)",
            "expected_results": "Generated expected results (implement parsing)"
        }












"""
Configuration file for Test Case Processor
"""

import os
from typing import List

class Config:
    # API Configuration
    OPENAI_API_KEY: str = os.getenv('OPENAI_API_KEY', 'your-openai-api-key-here')
    AI_MODEL: str = 'gpt-3.5-turbo'  # or 'gpt-4' for better quality
    MAX_TOKENS: int = 1000
    TEMPERATURE: float = 0.3  # Lower temperature for more consistent results
    API_DELAY: float = 1.0  # Delay between API calls in seconds
    
    # Excel Column Configuration
    REQUIRED_COLUMNS: List[str] = [
        'Test Case Description',
        'Test Case Steps'
    ]
    
    # Optional columns that will be preserved if they exist
    OPTIONAL_COLUMNS: List[str] = [
        'Test Case ID',
        'Test Case Name',
        'Priority',
        'Category',
        'Preconditions',
        'Test Data',
        'Expected Results'
    ]
    
    # Processing Configuration
    DEFAULT_BATCH_SIZE: int = 10
    MAX_RETRIES: int = 3
    RETRY_DELAY: float = 2.0
    
    # File Paths
    INPUT_FOLDER: str = 'input'
    OUTPUT_FOLDER: str = 'output'
    LOGS_FOLDER: str = 'logs'
    
    # Logging Configuration
    LOG_LEVEL: str = 'INFO'
    LOG_FORMAT: str = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    
    def __init__(self):
        """Initialize configuration and validate settings"""
        self.validate_config()
    
    def validate_config(self):
        """Validate configuration settings"""
        if not self.OPENAI_API_KEY or self.OPENAI_API_KEY == 'your-openai-api-key-here':
            print("⚠️  Warning: OpenAI API key not set. Please set OPENAI_API_KEY environment variable.")
            print("   You can also modify the config.py file directly.")
        
        # Create necessary directories
        for folder in [self.INPUT_FOLDER, self.OUTPUT_FOLDER, self.LOGS_FOLDER]:
            os.makedirs(folder, exist_ok=True)
    
    @classmethod
    def from_env(cls):
        """Create configuration from environment variables"""
        config = cls()
        
        # Override with environment variables if they exist
        config.OPENAI_API_KEY = os.getenv('OPENAI_API_KEY', config.OPENAI_API_KEY)
        config.AI_MODEL = os.getenv('AI_MODEL', config.AI_MODEL)
        config.MAX_TOKENS = int(os.getenv('MAX_TOKENS', config.MAX_TOKENS))
        config.TEMPERATURE = float(os.getenv('TEMPERATURE', config.TEMPERATURE))
        config.API_DELAY = float(os.getenv('API_DELAY', config.API_DELAY))
        config.DEFAULT_BATCH_SIZE = int(os.getenv('BATCH_SIZE', config.DEFAULT_BATCH_SIZE))
        
        return config









# Test Case Processor Requirements
pandas>=1.5.0
openpyxl>=3.0.0
openai>=1.0.0
requests>=2.28.0
python-dotenv>=1.0.0

# Optional: For advanced Excel processing
xlsxwriter>=3.0.0

# Optional: For local AI models
# ollama-python>=0.1.0

# Development dependencies (optional)
pytest>=7.0.0
black>=22.0.0
flake8>=6.0.0






#!/usr/bin/env python3
"""
Setup script for Test Case Processor
Creates folder structure and sample files
"""

import os
import pandas as pd
from pathlib import Path

def create_folder_structure():
    """Create necessary folders"""
    folders = [
        'input',
        'output',
        'logs'
    ]
    
    for folder in folders:
        Path(folder).mkdir(exist_ok=True)
        print(f"✅ Created folder: {folder}")

def create_sample_test_cases():
    """Create a sample Excel file with test cases"""
    sample_data = [
        {
            'Test Case ID': 'TC001',
            'Test Case Description': 'Login with valid credentials',
            'Test Case Steps': '1. Open login page 2. Enter username 3. Enter password 4. Click login button',
            'Priority': 'High',
            'Category': 'Authentication'
        },
        {
            'Test Case ID': 'TC002',
            'Test Case Description': 'Add item to shopping cart',
            'Test Case Steps': '1. Browse products 2. Select item 3. Click add to cart 4. Verify item added',
            'Priority': 'Medium',
            'Category': 'E-commerce'
        },
        {
            'Test Case ID': 'TC003',
            'Test Case Description': 'Search for products',
            'Test Case Steps': '1. Enter search term 2. Click search 3. Review results 4. Verify relevant products shown',
            'Priority': 'High',
            'Category': 'Search'
        },
        {
            'Test Case ID': 'TC004',
            'Test Case Description': 'User registration process',
            'Test Case Steps': '1. Go to signup page 2. Fill registration form 3. Submit form 4. Verify account created',
            'Priority': 'High',
            'Category': 'Authentication'
        },
        {
            'Test Case ID': 'TC005',
            'Test Case Description': 'Password reset functionality',
            'Test Case Steps': '1. Click forgot password 2. Enter email 3. Check email for reset link 4. Reset password',
            'Priority': 'Medium',
            'Category': 'Authentication'
        },
        {
            'Test Case ID': 'TC006',
            'Test Case Description': 'Checkout process',
            'Test Case Steps': '1. Add items to cart 2. Proceed to checkout 3. Enter shipping info 4. Complete payment',
            'Priority': 'High',
            'Category': 'E-commerce'
        },
        {
            'Test Case ID': 'TC007',
            'Test Case Description': 'Product filtering',
            'Test Case Steps': '1. Go to products page 2. Apply price filter 3. Apply category filter 4. Verify filtered results',
            'Priority': 'Medium',
            'Category': 'Search'
        },
        {
            'Test Case ID': 'TC008',
            'Test Case Description': 'User profile update',
            'Test Case Steps': '1. Login to account 2. Go to profile 3. Update information 4. Save changes',
            'Priority': 'Low',
            'Category': 'User Management'
        },
        {
            'Test Case ID': 'TC009',
            'Test Case Description': 'Order history review',
            'Test Case Steps': '1. Login to account 2. Navigate to order history 3. View past orders 4. Check order details',
            'Priority': 'Low',
            'Category': 'User Management'
        },
        {
            'Test Case ID': 'TC010',
            'Test Case Description': 'Contact form submission',
            'Test Case Steps': '1. Go to contact page 2. Fill contact form 3. Submit form 4. Verify confirmation message',
            'Priority': 'Medium',
            'Category': 'Communication'
        },
        {
            'Test Case ID': 'TC011',
            'Test Case Description': 'Mobile responsive design test',
            'Test Case Steps': '1. Open site on mobile 2. Test navigation 3. Test form inputs 4. Verify layout adaptation',
            'Priority': 'High',
            'Category': 'Responsive Design'
        },
        {
            'Test Case ID': 'TC012',
            'Test Case Description': 'API endpoint testing',
            'Test Case Steps': '1. Send GET request 2. Verify response status 3. Check response data 4. Validate JSON structure',
            'Priority': 'High',
            'Category': 'API Testing'
        }
    ]
    
    df = pd.DataFrame(sample_data)
    sample_file = 'input/test_cases.xlsx'
    df.to_excel(sample_file, index=False)
    print(f"✅ Created sample file: {sample_file}")
    print(f"   📊 Sample contains {len(sample_data)} test cases")

def create_env_file():
    """Create .env file template"""
    env_content = """# Test Case Processor Environment Variables

# OpenAI API Configuration
OPENAI_API_KEY=your-openai-api-key-here
AI_MODEL=gpt-3.5-turbo
MAX_TOKENS=1000
TEMPERATURE=0.3

# Processing Configuration  
BATCH_SIZE=10
API_DELAY=1.0

# Optional: Local AI Configuration (if using Ollama)
# LOCAL_AI_URL=http://localhost:11434
# LOCAL_AI_MODEL=llama2
"""
    
    with open('.env', 'w') as f:
        f.write(env_content)
    print("✅ Created .env template file")
    print("   ⚠️  Please update .env with your actual API keys")

def create_readme():
    """Create README file"""
    readme_content = """# Test Case Processor

An AI-powered tool to rephrase and enhance Excel test cases.

## Setup

1. Install dependencies:
   ```bash
   pip install -r requirements.txt
   ```

2. Set up your OpenAI API key:
   - Update the `.env` file with your API key
   - Or set environment variable: `export OPENAI_API_KEY=your-key-here`

3. Place your test case Excel file in the `input/` folder

## Usage

### Basic Usage
```python
from main import TestCaseProcessor

processor = TestCaseProcessor()
processor.process_file('input/your_test_cases.xlsx')
```

### Command Line Usage
```bash
python main.py
```

## Folder Structure
```
test-case-processor/
├── input/           # Place your Excel files here
├── output/          # Processed files will be saved here  
├── logs/            # Log files
├── main.py          # Main processing script
├── test_case_ai.py  # AI processing module
├── config.py        # Configuration settings
├── requirements.txt # Python dependencies
└── .env            # Environment variables
```

## Configuration

Edit `config.py` to customize:
- AI model settings
- Column mappings
- Batch processing size
- API rate limits

## Sample Data

A sample Excel file with 12 test cases is provided in `input/test_cases.xlsx` for testing.

## Features

- ✅ Batch processing (10+ test cases at once)
- ✅ AI-powered rephrasing using OpenAI GPT
- ✅ Preserves all original columns
- ✅ Generates expected results automatically
- ✅ Error handling and logging
- ✅ Rate limiting for API calls
- ✅ Configurable processing parameters

## Requirements

- Python 3.7+
- OpenAI API key
- Excel files with columns: "Test Case Description", "Test Case Steps"
"""
    
    with open('README.md', 'w') as f:
        f.write(readme_content)
    print("✅ Created README.md")

def main():
    """Run setup"""
    print("🚀 Setting up Test Case Processor...")
    print("=" * 50)
    
    create_folder_structure()
    create_sample_test_cases()
    create_env_file()
    create_readme()
    
    print("\n" + "=" * 50)
    print("✅ Setup complete!")
    print("\n📋 Next steps:")
    print("1. Install dependencies: pip install -r requirements.txt")
    print("2. Update .env file with your OpenAI API key")
    print("3. Run: python main.py")
    print("\n💡 A sample Excel file has been created in input/test_cases.xlsx")

if __name__ == "__main__":
    main()









#!/usr/bin/env python3
"""
Example usage script for Test Case Processor
Demonstrates different ways to use the processor
"""

from main import TestCaseProcessor
from test_case_ai import TestCaseAI
import pandas as pd
import os

def example_basic_processing():
    """Example 1: Basic file processing"""
    print("📝 Example 1: Basic Processing")
    print("-" * 40)
    
    processor = TestCaseProcessor()
    
    # Process the sample file
    input_file = "input/test_cases.xlsx"
    
    if os.path.exists(input_file):
        result_df = processor.process_file(
            input_file=input_file,
            batch_size=5  # Process 5 test cases at a time
        )
        print(f"✅ Processed {len(result_df)} test cases")
        print(f"📁 Output saved to: output/test_cases_rephrased.xlsx")
    else:
        print(f"❌ Input file not found: {input_file}")
        print("   Run 'python setup.py' to create sample data")

def example_custom_output():
    """Example 2: Custom output file"""
    print("\n📝 Example 2: Custom Output File")
    print("-" * 40)
    
    processor = TestCaseProcessor()
    
    input_file = "input/test_cases.xlsx"
    custom_output = "output/my_enhanced_test_cases.xlsx"
    
    if os.path.exists(input_file):
        result_df = processor.process_file(
            input_file=input_file,
            output_file=custom_output,
            batch_size=10
        )
        print(f"✅ Processed {len(result_df)} test cases")
        print(f"📁 Custom output saved to: {custom_output}")
    else:
        print(f"❌ Input file not found: {input_file}")

def example_single_test_case():
    """Example 3: Process a single test case"""
    print("\n📝 Example 3: Single Test Case Processing")
    print("-" * 40)
    
    ai_processor = TestCaseAI()
    
    # Test API connection first
    if not ai_processor.test_connection():
        print("❌ AI API connection failed. Check your API key.")
        return
    
    # Sample test case
    description = "Login with valid credentials"
    steps = "1. Open login page 2. Enter username 3. Enter password 4. Click login"
    
    print(f"Original Description: {description}")
    print(f"Original Steps: {steps}")
    print("\n🤖 Processing with AI...")
    
    try:
        result = ai_processor.rephrase_test_case(description, steps)
        
        print("\n✅ AI-Enhanced Test Case:")
        print(f"Description: {result['description']}")
        print(f"Steps: {result['steps']}")
        print(f"Expected Results: {result['expected_results']}")
        
    except Exception as e:
        print(f"❌ Processing failed: {e}")

def example_batch_analysis():
    """Example 4: Analyze processed results"""
    print("\n📝 Example 4: Batch Analysis")
    print("-" * 40)
    
    output_file = "output/test_cases_rephrased.xlsx"
    
    if os.path.exists(output_file):
        df = pd.read_excel(output_file)
        
        print(f"📊 Analysis of processed file:")
        print(f"   Total test cases: {len(df)}")
        
        if 'AI_Processed' in df.columns:
            successful = df['AI_Processed'].sum()
            failed = len(df) - successful
            print(f"   Successfully processed: {successful}")
            print(f"   Failed: {failed}")
        
        if 'Category' in df.columns:
            print(f"\n📈 Test cases by category:")
            category_counts = df['Category'].value_counts()
            for category, count in category_counts.items():
                print(f"   {category}: {count}")
        
        if 'Priority' in df.columns:
            print(f"\n⚡ Test cases by priority:")
            priority_counts = df['Priority'].value_counts()
            for priority, count in priority_counts.items():
                print(f"   {priority}: {count}")
        
        # Show sample of enhanced content
        if len(df) > 0:
            print(f"\n📋 Sample enhanced test case:")
            sample = df.iloc[0]
            print(f"   ID: {sample.get('Test Case ID', 'N/A')}")
            print(f"   Original vs Enhanced:")
            print(f"   Description: {sample.get('Test Case Description', '')[:100]}...")
            if 'Expected Results' in df.columns:
                print(f"   Expected Results: {sample.get('Expected Results', '')[:100]}...")
    
    else:
        print(f"❌ Output file not found: {output_file}")
        print("   Run example_basic_processing() first")

def example_error_handling():
    """Example 5: Error handling and recovery"""
    print("\n📝 Example 5: Error Handling Demo")
    print("-" * 40)
    
    # Create test data with potential issues
    test_data = [
        {
            'Test Case ID': 'TC_ERROR1',
            'Test Case Description': '',  # Empty description
            'Test Case Steps': 'Step 1: Do something',
            'Priority': 'High'
        },
        {
            'Test Case ID': 'TC_ERROR2', 
            'Test Case Description': 'Test with very long description that might cause issues with token limits in the AI processing pipeline and should be handled gracefully by the system',
            'Test Case Steps': '',  # Empty steps
            'Priority': 'Medium'
        }
    ]
    
    # Create temporary test file
    test_df = pd.DataFrame(test_data)
    temp_file = "input/error_test.xlsx"
    test_df.to_excel(temp_file, index=False)
    
    print(f"📄 Created test file with problematic data: {temp_file}")
    
    processor = TestCaseProcessor()
    
    try:
        result_df = processor.process_file(
            input_file=temp_file,
            output_file="output/error_test_results.xlsx",
            batch_size=2
        )
        
        # Analyze results
        if 'AI_Processed' in result_df.columns:
            successful = result_df['AI_Processed'].sum()
            failed = len(result_df) - successful
            print(f"✅ Error handling test complete:")
            print(f"   Successful: {successful}")
            print(f"   Failed: {failed}")
            
            if failed > 0:
                failed_cases = result_df[result_df['AI_Processed'] == False]
                print(f"   Failed cases: {list(failed_cases['Test Case ID'])}")
    
    except Exception as e:
        print(f"❌ Error handling test failed: {e}")
    
    finally:
        # Clean up
        if os.path.exists(temp_file):
            os.remove(temp_file)
            print(f"🧹 Cleaned up temporary file: {temp_file}")

def main():
    """Run all examples"""
    print("🚀 Test Case Processor - Example Usage")
    print("=" * 50)
    
    # Check if setup has been run
    if not os.path.exists("input"):
        print("❌ Setup not complete. Please run 'python setup.py' first.")
        return
    
    # Run examples
    example_basic_processing()
    example_custom_output()
    example_single_test_case()
    example_batch_analysis()
    example_error_handling()
    
    print("\n" + "=" * 50)
    print("✅ All examples completed!")
    print("\n💡 Tips:")
    print("- Check the 'output/' folder for processed files")
    print("- Review 'logs/' folder for detailed processing logs")
    print("- Modify 'config.py' to customize AI behavior")
    print("- Set OPENAI_API_KEY environment variable for AI processing")

if __name__ == "__main__":
    main()










#!/usr/bin/env python3
"""
Format-Preserving Excel Processor
Maintains original Excel formatting while updating content
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import logging
from typing import Dict, Any, List
from copy import copy
import os

logger = logging.getLogger(__name__)

class FormatPreservingProcessor:
    def __init__(self):
        self.original_styles = {}
        self.column_mappings = {}
        
    def load_workbook_with_formatting(self, file_path: str):
        """Load workbook while preserving all formatting"""
        try:
            # Load with openpyxl to preserve formatting
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            worksheet = workbook.active
            
            logger.info(f"Loaded workbook with {worksheet.max_row} rows and {worksheet.max_column} columns")
            return workbook, worksheet
            
        except Exception as e:
            logger.error(f"Error loading workbook: {e}")
            raise
    
    def extract_formatting(self, worksheet):
        """Extract all formatting information from the worksheet"""
        formatting_data = {}
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.coordinate not in formatting_data:
                    formatting_data[cell.coordinate] = {
                        'font': copy(cell.font) if cell.font else None,
                        'fill': copy(cell.fill) if cell.fill else None,
                        'border': copy(cell.border) if cell.border else None,
                        'alignment': copy(cell.alignment) if cell.alignment else None,
                        'number_format': cell.number_format,
                        'protection': copy(cell.protection) if cell.protection else None
                    }
        
        # Also preserve row heights and column widths
        row_dimensions = {}
        for row_num in range(1, worksheet.max_row + 1):
            if worksheet.row_dimensions[row_num].height:
                row_dimensions[row_num] = worksheet.row_dimensions[row_num].height
        
        column_dimensions = {}
        for col_letter in [worksheet.cell(1, col).column_letter for col in range(1, worksheet.max_column + 1)]:
            if worksheet.column_dimensions[col_letter].width:
                column_dimensions[col_letter] = worksheet.column_dimensions[col_letter].width
        
        return formatting_data, row_dimensions, column_dimensions
    
    def apply_formatting(self, worksheet, formatting_data, row_dimensions, column_dimensions):
        """Apply preserved formatting to the worksheet"""
        # Apply cell formatting
        for coordinate, formats in formatting_data.items():
            try:
                cell = worksheet[coordinate]
                if formats['font']:
                    cell.font = formats['font']
                if formats['fill']:
                    cell.fill = formats['fill']
                if formats['border']:
                    cell.border = formats['border']
                if formats['alignment']:
                    cell.alignment = formats['alignment']
                if formats['number_format']:
                    cell.number_format = formats['number_format']
                if formats['protection']:
                    cell.protection = formats['protection']
            except Exception as e:
                logger.warning(f"Could not apply formatting to {coordinate}: {e}")
        
        # Apply row heights
        for row_num, height in row_dimensions.items():
            worksheet.row_dimensions[row_num].height = height
        
        # Apply column widths
        for col_letter, width in column_dimensions.items():
            worksheet.column_dimensions[col_letter].width = width
    
    def get_column_mapping(self, worksheet):
        """Get mapping of column names to column indices"""
        headers = {}
        header_row = 1  # Assume first row contains headers
        
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(header_row, col).value
            if cell_value:
                headers[str(cell_value).strip()] = col
        
        logger.info(f"Found columns: {list(headers.keys())}")
        return headers
    
    def convert_to_dataframe(self, worksheet):
        """Convert worksheet to DataFrame while preserving data"""
        data = []
        headers = []
        
        # Get headers from first row
        for col in range(1, worksheet.max_column + 1):
            header_cell = worksheet.cell(1, col)
            headers.append(header_cell.value if header_cell.value else f"Column_{col}")
        
        # Get data from remaining rows
        for row in range(2, worksheet.max_row + 1):
            row_data = []
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row, col)
                row_data.append(cell.value)
            data.append(row_data)
        
        df = pd.DataFrame(data, columns=headers)
        return df
    
    def update_worksheet_with_dataframe(self, worksheet, df, start_row=2):
        """Update worksheet with DataFrame data while preserving formatting"""
        # Clear existing data (keep headers)
        for row in range(start_row, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row, col).value = None
        
        # Add new data
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=start_row):
            for col_idx, (col_name, value) in enumerate(row_data.items(), start=1):
                cell = worksheet.cell(row_idx, col_idx)
                cell.value = value
    
    def process_with_format_preservation(self, input_file: str, output_file: str, 
                                       processed_df: pd.DataFrame):
        """Main function to process while preserving Excel formatting"""
        try:
            logger.info("Loading original workbook with formatting...")
            workbook, worksheet = self.load_workbook_with_formatting(input_file)
            
            logger.info("Extracting formatting information...")
            formatting_data, row_dims, col_dims = self.extract_formatting(worksheet)
            
            logger.info("Getting column mappings...")
            column_mapping = self.get_column_mapping(worksheet)
            
            logger.info("Converting original data to DataFrame...")
            original_df = self.convert_to_dataframe(worksheet)
            
            logger.info("Updating worksheet with processed data...")
            self.update_worksheet_with_dataframe(worksheet, processed_df)
            
            logger.info("Reapplying formatting...")
            self.apply_formatting(worksheet, formatting_data, row_dims, col_dims)
            
            # Handle additional rows if processed_df has more rows than original
            if len(processed_df) > len(original_df):
                logger.info("Extending formatting for additional rows...")
                self.extend_formatting_for_new_rows(worksheet, formatting_data, 
                                                  len(original_df), len(processed_df))
            
            logger.info(f"Saving formatted workbook to {output_file}...")
            # Create output directory if it doesn't exist
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            workbook.save(output_file)
            
            logger.info("✅ Successfully saved with preserved formatting!")
            
        except Exception as e:
            logger.error(f"Error in format preservation: {e}")
            raise
    
    def extend_formatting_for_new_rows(self, worksheet, formatting_data, 
                                     original_rows: int, new_rows: int):
        """Extend formatting for newly added rows"""
        # Get formatting from the last data row
        last_data_row = original_rows + 1  # +1 because we start from row 2 (after headers)
        
        for new_row in range(original_rows + 2, new_rows + 2):  # +2 for header offset
            for col in range(1, worksheet.max_column + 1):
                # Copy formatting from last data row
                source_coordinate = f"{worksheet.cell(last_data_row, col).column_letter}{last_data_row}"
                target_coordinate = f"{worksheet.cell(new_row, col).column_letter}{new_row}"
                
                if source_coordinate in formatting_data:
                    source_formats = formatting_data[source_coordinate]
                    target_cell = worksheet[target_coordinate]
                    
                    if source_formats['font']:
                        target_cell.font = copy(source_formats['font'])
                    if source_formats['fill']:
                        target_cell.fill = copy(source_formats['fill'])
                    if source_formats['border']:
                        target_cell.border = copy(source_formats['border'])
                    if source_formats['alignment']:
                        target_cell.alignment = copy(source_formats['alignment'])
                    if source_formats['number_format']:
                        target_cell.number_format = source_formats['number_format']

# Enhanced main processor that uses format preservation
class EnhancedTestCaseProcessor:
    def __init__(self):
        from test_case_ai import TestCaseAI
        from config import Config
        
        self.config = Config()
        self.ai_processor = TestCaseAI()
        self.format_processor = FormatPreservingProcessor()
        
    def process_file_with_formatting(self, input_file: str, output_file: str = None, 
                                   batch_size: int = 10):
        """Process test cases while maintaining Excel formatting"""
        try:
            # First, load the file normally to get data for processing
            logger.info("Loading data for AI processing...")
            df = pd.read_excel(input_file)
            
            # Validate columns
            required_columns = ['Test Case Description', 'Test Case Steps']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                raise ValueError(f"Missing required columns: {missing_columns}")
            
            # Process data with AI (same as before)
            logger.info("Processing with AI...")
            processed_df = self.process_dataframe_with_ai(df, batch_size)
            
            # Set output file name if not provided
            if output_file is None:
                base_name = os.path.splitext(os.path.basename(input_file))[0]
                output_file = f"output/{base_name}_formatted_enhanced.xlsx"
            
            # Now use format preservation to save
            logger.info("Applying format preservation...")
            self.format_processor.process_with_format_preservation(
                input_file, output_file, processed_df
            )
            
            return processed_df
            
        except Exception as e:
            logger.error(f"Enhanced processing failed: {e}")
            raise
    
    def process_dataframe_with_ai(self, df: pd.DataFrame, batch_size: int = 10):
        """Process DataFrame with AI enhancement"""
        processed_df = df.copy()
        
        # Add new columns if they don't exist
        if 'Expected Results' not in processed_df.columns:
            processed_df['Expected Results'] = ''
        if 'AI_Processed' not in processed_df.columns:
            processed_df['AI_Processed'] = False
        if 'Processing_Timestamp' not in processed_df.columns:
            processed_df['Processing_Timestamp'] = ''
        
        total_rows = len(df)
        
        for start_idx in range(0, total_rows, batch_size):
            end_idx = min(start_idx + batch_size, total_rows)
            logger.info(f"Processing batch: rows {start_idx} to {end_idx-1}")
            
            for idx in range(start_idx, end_idx):
                try:
                    original_description = str(processed_df.iloc[idx]['Test Case Description'])
                    original_steps = str(processed_df.iloc[idx]['Test Case Steps'])
                    
                    # Generate rephrased content using AI
                    rephrased_data = self.ai_processor.rephrase_test_case(
                        description=original_description,
                        steps=original_steps
                    )
                    
                    # Update the DataFrame
                    processed_df.iloc[idx, processed_df.columns.get_loc('Test Case Description')] = rephrased_data['description']
                    processed_df.iloc[idx, processed_df.columns.get_loc('Test Case Steps')] = rephrased_data['steps']
                    processed_df.iloc[idx, processed_df.columns.get_loc('Expected Results')] = rephrased_data['expected_results']
                    processed_df.iloc[idx, processed_df.columns.get_loc('AI_Processed')] = True
                    processed_df.iloc[idx, processed_df.columns.get_loc('Processing_Timestamp')] = pd.Timestamp.now()
                    
                    # Rate limiting
                    import time
                    time.sleep(self.config.API_DELAY)
                    
                except Exception as e:
                    logger.error(f"Error processing row {idx}: {e}")
                    processed_df.iloc[idx, processed_df.columns.get_loc('AI_Processed')] = False
        
        return processed_df

def main():
    """Example usage of format-preserving processor"""
    processor = EnhancedTestCaseProcessor()
    
    input_file = "input/test_cases.xlsx"
    
    try:
        result_df = processor.process_file_with_formatting(
            input_file=input_file,
            batch_size=10
        )
        print("✅ Processing completed with preserved formatting!")
        print(f"📊 Processed {len(result_df)} test cases")
        
    except Exception as e:
        print(f"❌ Processing failed: {e}")

if __name__ == "__main__":
    main()



